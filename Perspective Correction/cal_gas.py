import cv2
import numpy as np
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter
from io import BytesIO
from tqdm import tqdm
import PIL.Image as PILImage


def setup_environment():
    """初始化环境"""
    import warnings
    warnings.filterwarnings("ignore")


def get_target_points(min_x, min_y, target_width, target_height):
    """生成目标矩形的四个角点"""
    return np.float32([
        [min_x, min_y],
        [min_x, min_y + target_height],
        [min_x + target_width, min_y + target_height],
        [min_x + target_width, min_y]
    ])


def perform_perspective_correction(image, points, output_path=None):
    """执行透视变换矫正"""
    min_x, max_x = np.min(points[:, 0]), np.max(points[:, 0])
    min_y, max_y = np.min(points[:, 1]), np.max(points[:, 1])
    target_width = int(max_x - min_x)
    target_height = int(max_y - min_y)
    dst_pts = get_target_points(min_x, min_y, target_width, target_height)

    matrix = cv2.getPerspectiveTransform(points, dst_pts)
    corrected = cv2.warpPerspective(image, matrix, (image.shape[1], image.shape[0]))

    if output_path:
        cv2.imwrite(output_path, corrected)
    return corrected, matrix


def analyze_geometric_alignment(img):
    """
    专业几何对齐度分析（增强版）
    :return: {
        'vertical_score': 垂直对齐度 (0-1),
        'horizontal_score': 水平对齐度 (0-1),
        'total_score': 综合几何对齐度 (0-1)
    }
    """
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

    # 增强的边缘检测
    edges = cv2.Canny(gray, 50, 150, apertureSize=3)

    # 使用概率霍夫变换检测更精确的直线
    lines = cv2.HoughLinesP(edges, 1, np.pi / 180, threshold=50,
                            minLineLength=30, maxLineGap=10)

    if lines is None:
        return {'vertical_score': 0, 'horizontal_score': 0, 'total_score': 0}

    vertical_angles = []
    horizontal_angles = []

    for line in lines:
        x1, y1, x2, y2 = line[0]
        angle = np.arctan2(y2 - y1, x2 - x1) * 180 / np.pi

        # 角度归一化到[-90,90]
        if angle > 90:
            angle -= 180
        if angle < -90:
            angle += 180

        # 分类为垂直或水平
        if abs(angle) < 15:  # 水平线（-15°到15°）
            horizontal_angles.append(abs(angle))
        elif abs(angle - 90) < 15:  # 垂直线（75°到105°）
            vertical_angles.append(abs(angle - 90))

    # 计算垂直对齐度（越接近90°越好）
    vertical_score = 1 - np.mean(vertical_angles) / 15 if vertical_angles else 0

    # 计算水平对齐度（越接近0°越好）
    horizontal_score = 1 - np.mean(horizontal_angles) / 15 if horizontal_angles else 0

    # 综合评分（加权平均）
    total_score = 0.6 * vertical_score + 0.4 * horizontal_score

    return {
        'total_score': min(max(total_score, 0), 1),
        'vertical_lines': len(vertical_angles),
        'horizontal_lines': len(horizontal_angles)
    }


def prepare_image_for_excel(image, max_size=300):
    """准备用于Excel的图片"""
    # 调整大小
    h, w = image.shape[:2]
    scale = min(max_size / w, max_size / h)
    resized = cv2.resize(image, (int(w * scale), int(h * scale)))

    # 转换为RGB
    resized = cv2.cvtColor(resized, cv2.COLOR_BGR2RGB)

    # 转换为PIL Image
    pil_img = PILImage.fromarray(resized)

    # 保存到字节流
    img_bytes = BytesIO()
    pil_img.save(img_bytes, format='JPEG')
    img_bytes.seek(0)

    return img_bytes


def process_image_for_excel(image_path, points_path1, points_path2):
    """处理单张图像并返回所有需要的数据"""
    image = cv2.imread(image_path)
    if image is None: return None

    # 读取角点
    points1 = np.loadtxt(points_path1, dtype=np.float32)
    points2 = np.loadtxt(points_path2, dtype=np.float32)
    if points1.shape != (4, 2) or points2.shape != (4, 2): return None

    # 执行矫正
    corrected1, _ = perform_perspective_correction(image.copy(), points1)
    corrected2, _ = perform_perspective_correction(image.copy(), points2)

    # 分析几何对齐度
    algo_score = analyze_geometric_alignment(corrected1)
    manual_score = analyze_geometric_alignment(corrected2)

    # 准备图片数据
    original_img_bytes = prepare_image_for_excel(image)
    algo_img_bytes = prepare_image_for_excel(corrected1)
    manual_img_bytes = prepare_image_for_excel(corrected2)

    return {
        'filename': os.path.basename(image_path),
        'original_image': original_img_bytes,
        'algorithm_image': algo_img_bytes,
        'manual_image': manual_img_bytes,
        'algo_score': algo_score['total_score'],
        'manual_score': manual_score['total_score'],
        'score_diff': abs(algo_score['total_score'] - manual_score['total_score']),
        'algo_vertical': algo_score['vertical_lines'],
        'manual_vertical': manual_score['vertical_lines'],
        'algo_horizontal': algo_score['horizontal_lines'],
        'manual_horizontal': manual_score['horizontal_lines']
    }


def generate_excel_report(results, output_path):
    """生成包含图片的Excel报告"""
    wb = Workbook()
    ws = wb.active
    ws.title = "矫正结果对比"

    # 设置表头
    headers = ["文件名", "原图", "算法矫正结果", "人工矫正结果",
               "算法评分", "人工评分", "评分差值",
               "算法垂直线", "人工垂直线", "算法水平线", "人工水平线"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # 写入数据并插入图片
    for row_idx, result in enumerate(results, 2):
        # 写入文本数据
        ws.cell(row=row_idx, column=1, value=result['filename'])
        ws.cell(row=row_idx, column=5, value=result['algo_score'])
        ws.cell(row=row_idx, column=6, value=result['manual_score'])
        ws.cell(row=row_idx, column=7, value=result['score_diff'])
        ws.cell(row=row_idx, column=8, value=result['algo_vertical'])
        ws.cell(row=row_idx, column=9, value=result['manual_vertical'])
        ws.cell(row=row_idx, column=10, value=result['algo_horizontal'])
        ws.cell(row=row_idx, column=11, value=result['manual_horizontal'])

        # 插入图片
        def add_image(col, img_bytes):
            img = ExcelImage(img_bytes)
            img.anchor = f"{get_column_letter(col)}{row_idx}"
            ws.add_image(img)

        add_image(2, result['original_image'])
        add_image(3, result['algorithm_image'])
        add_image(4, result['manual_image'])

    # 调整列宽和行高
    for col in range(1, 12):
        if col in [2, 3, 4]:  # 图片列
            ws.column_dimensions[get_column_letter(col)].width = 30
        else:  # 数据列
            ws.column_dimensions[get_column_letter(col)].width = 15

    for row in range(2, len(results) + 2):
        ws.row_dimensions[row].height = 150

    wb.save(output_path)
    print(f"Excel报告已生成: {output_path}")


def main():
    """主函数"""
    setup_environment()

    # 配置路径
    image_dir = r'./angle_adjust_train_ori'
    algo_points_dir = r'./points-algorithmic'
    manual_points_dir = r'./points-manual'
    output_excel = './矫正结果对比报告.xlsx'

    # 处理所有图像
    results = []
    for filename in tqdm(os.listdir(image_dir)):
        if not filename.lower().endswith(('.jpg', '.jpeg', '.png', '.bmp')):
            continue

        base_name = os.path.splitext(filename)[0]
        image_path = os.path.join(image_dir, filename)
        points_path1 = os.path.join(algo_points_dir, f"{base_name}.txt")
        points_path2 = os.path.join(manual_points_dir, f"{base_name}_point.txt")

        result = process_image_for_excel(image_path, points_path1, points_path2)
        if result:
            results.append(result)
            print(f"处理完成: {filename} | 算法评分: {result['algo_score']} | 人工评分: {result['manual_score']}")

    # 生成Excel报告
    if results:
        generate_excel_report(results, output_excel)
        print(f"\n成功处理 {len(results)} 张图片，报告已保存至: {output_excel}")
    else:
        print("没有成功处理任何图像")


if __name__ == "__main__":
    main()